import openpyxl
from openpyxl.styles import Font
from config.config_global import *
import time


# from utils.Log import *


class OperationExcel(object):
    '''
        定义了两个全局变量file_name和sheet_id,
        判断file_name如果为None就给个默认的值
    '''

    def __init__(self, file_name=None, sheet_id=None):

        self.workbook = None
        self.excelFile = excelpath
        self.font = Font(color=None)  # 设置字体的颜色
        # 颜色对应的RGB值
        self.RGBDict = {'red': 'FFFF3030', 'green': 'FF008B00', 'yellow': 'E2FF00'}
        if file_name:
            self.file_name = file_name
            self.sheet_id = sheet_id

        else:
            self.file_name = excelpath
            self.sheet_id = 0

        self.load_workBook = self.loadWorkBook(excelpath)
        self.data = self.get_data()

    # 将excel文件加载到内存，并获取其workbook对象
    def loadWorkBook(self, excelPathAndName):

        try:
            self.workbook = openpyxl.load_workbook(excelPathAndName)
        except Exception as e:
            pass

        self.excelFile = excelPathAndName
        return self.workbook

    '''获取excel的表'''

    def get_data(self):
        try:
            sheet = self.workbook["接口用例"]
            # sheet = self.workbook.sheetnames   #获取所有表名

            return sheet
        except Exception as e:

            pass

    '''获取sheet表的行数'''

    def get_lines(self):
        tables = self.data
        return tables.max_row

    '''获取某一个单元格里的值'''

    def get_cell_value(self, row, col):
        try:
            return self.data.cell(row=row, column=col).value
        except Exception as e:
            # raise e
            pass

    '''往excel里写入数据'''

    def write_value(self, sheet, content, coordinate=None, rowNo=None, colsNo=None, style=None):
        if coordinate is not None:  # 新版本报错
            try:
                sheet.cell(coordinate=coordinate).value = content
                if style is not None:
                    sheet.cell(coordinate=coordinate).font = Font(color=self.RGBDict[style])
                self.workbook.save(self.excelFile)
            except Exception as e:
                pass
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row=rowNo, column=colsNo).value = content
                if style:
                    sheet.cell(row=rowNo, column=colsNo).font = Font(color=self.RGBDict[style])
                self.workbook.save(excelpath)
            except Exception as e:
                pass
        else:
            raise Exception("Insufficient Coordinates od cell。。。")

        '''
           写入当前的时间，下标从1开始
        '''

    def writeCellCurrentTime(self, sheet, coordinate=None, rowNo=None, colsNo=None):
        now = int(time.time())
        timeArray = time.localtime(now)
        currentTime = time.strftime("%Y-%m-%d %H:%M:%S", timeArray)

        if coordinate is not None:  # 新版本报错
            try:
                sheet.cell(coordinate=coordinate).value = currentTime
                self.workbook.save(self.excelFile)
            except Exception as e:
                # raise e
                pass
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row=rowNo, column=colsNo).value = currentTime
                self.workbook.save(self.excelFile)
            except Exception as e:
                pass
        else:
            raise Exception("Insufficient Coordinates od cell。。。")

    '''根据对应的case_id找到对应行的内容'''

    def get_rows_data(self, case_id):
        try:
            row_num = self.get_row_num(case_id)
            rows_data = self.get_row_values(row_num)
            return rows_data
        except Exception as e:
            pass

    '''根据对应的case_id找到对应的行号'''

    def get_row_num(self, case_id):
        try:
            num = 1
            clols_data = self.get_cols_data()
            for col_data in clols_data:
                if case_id in col_data.value:
                    return num
                num = num + 1
        except Exception as e:
            pass

    '''根据行号，找到该行的内容'''

    def get_row_values(self, row):
        try:
            tables = self.data
            row_data = list(tables.rows)[row - 1]
            return row_data
        except Exception as e:
            pass

    '''某一列的内容'''

    def get_cols_data(self, col_id=None):
        try:
            if col_id != None:
                cols = list(self.data.columns)[col_id - 1]
            else:
                cols = list(self.data.columns)[0]
            return cols
        except Exception as e:
            pass

    '''获取某行的内容，用于ddt'''

    def get_rows_value(self, row):
        row_list = []
        tables = self.data
        try:
            for i in list(tables.rows)[row - 1]:
                row_list.append(i.value)
            return row_list
        except Exception as e:
            pass

    def get_excel_all_data(self):
        '''
        获取excel里所有的数据
        :return:
        '''
        data_list = []
        for i in range(1, self.get_lines()):
            data_list.append(self.get_rows_value(i + 1))

        return data_list


if __name__ == '__main__':
    ob = OperationExcel()
    ob.loadWorkBook(excelpath)
    # print(ob.get_row_values(2))
    # for i in data:
    #     print(i.value)
    print(ob.get_excel_all_data())
    # logging.info(ob.get_data())
    # logging.info(ob.get_lines())
    # logging.info(ob.get_cell_value(2,2))
    # logging.info(ob.get_cols_data())
    # logging.info('fail')
    # logging.info(ob.get_row_num("login_005"))
    # logging.info(ob.get_row_values(ob.get_row_num("login_005")))
    # ob.write_value(ob.get_data(),content="oihh888g",rowNo=15,colsNo=8)

    # logging.info(ob.getStartRowNumber(sh))
    # logging.info(ob.getStartColNumber(sh))
    # logging.info(ob.getRow(sh,2))
    # logging.info(ob.getColumn(sh,2))
    # logging.info(ob.getCellOfValue(sh,rowNo=2,colsNo=2))
    # # logging.info(ob.getCellOfValue(sh,coordinate="B2"))
    # logging.info(ob.getCellOfObject(sh,rowNo=1,colsNo=2))
    # # logging.info(ob.writeCell(sh,'hello',coordinate='A2'))
    # logging.info(ob.writeCellCurrentTime(sh,rowNo=1,colsNo=2))
